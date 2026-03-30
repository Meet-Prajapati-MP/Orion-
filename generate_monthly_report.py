import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

OUTPUT = r"M:\Learning FrontEnd\orion\Monthly_Report_Feb_Mar_2026.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY        = "0D1B2A"
GOLD        = "E8C547"
WHITE       = "FFFFFF"
LIGHT_BG    = "F7F8FA"
ALT_ROW     = "EEF2FF"
GREEN_BG    = "D1FAE5"
GREEN_TXT   = "065F46"
RED_BG      = "FEE2E2"
RED_TXT     = "991B1B"
BLUE_BG     = "DBEAFE"
BLUE_TXT    = "1E40AF"
PURPLE_BG   = "EDE9FE"
PURPLE_TXT  = "5B21B6"
ORANGE_BG   = "FEF3C7"
ORANGE_TXT  = "92400E"
MID_GREY    = "6B7280"
DARK_TXT    = "1F2937"
BORDER_CLR  = "D1D5DB"
SECTION_BG  = "1E3A5F"

wb = openpyxl.Workbook()

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=DARK_TXT, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def border(style="thin", color=BORDER_CLR):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, row, col, value, bold=False, size=10, fcolor=None, tcolor=DARK_TXT,
         halign="left", wrap=False, italic=False, border_on=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=tcolor, italic=italic)
    c.alignment = align(h=halign, wrap=wrap)
    if fcolor:
        c.fill = fill(fcolor)
    if border_on:
        c.border = border()
    return c

def merge_cell(ws, r1, c1, r2, c2, value, bold=False, size=12, fcolor=NAVY,
               tcolor=WHITE, halign="center", wrap=False, italic=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = font(bold=bold, size=size, color=tcolor, italic=italic)
    c.fill = fill(fcolor)
    c.alignment = align(h=halign, v="center", wrap=wrap)
    c.border = border(style="medium", color=GOLD)
    return c

def section_header(ws, row, col, col_end, text):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=10, color=WHITE)
    c.fill = fill(SECTION_BG)
    c.alignment = align(h="left")
    c.border = border(style="medium", color=GOLD)

def col_header(ws, row, cols_vals, fcolor=NAVY, tcolor=WHITE, size=9):
    for col, val in cols_vals:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=True, size=size, color=tcolor)
        c.fill = fill(fcolor)
        c.alignment = align(h="center")
        c.border = border(style="medium", color=GOLD)

def stat_box(ws, row, col, label, value, unit="", fcolor=NAVY, tcolor=WHITE,
             val_color=GOLD):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = font(bold=True, size=9, color=WHITE)
    lc.fill = fill(fcolor)
    lc.alignment = align(h="center")
    lc.border = border(style="medium", color=GOLD)

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    vc = ws.cell(row=row+1, column=col, value=f"{value} {unit}".strip())
    vc.font = font(bold=True, size=16, color=val_color)
    vc.fill = fill(fcolor)
    vc.alignment = align(h="center")
    vc.border = border(style="medium", color=GOLD)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD / OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Dashboard"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A8"

# Title banner
merge_cell(ws1, 1, 1, 3, 12, "MONTHLY WORK REPORT  ·  February & March 2026",
           bold=True, size=18, fcolor=NAVY, tcolor=GOLD, halign="center")
merge_cell(ws1, 4, 1, 4, 12, "Intern: Dellc  ·  Projects: Curiosity Hub (InvoiceBill)  ·  Orion AI Multi-Agent Platform",
           bold=False, size=10, fcolor=SECTION_BG, tcolor=WHITE, halign="center")
merge_cell(ws1, 5, 1, 5, 12, "Period: 01 February 2026 — 27 March 2026  ·  Generated: 27 March 2026",
           bold=False, size=9, fcolor=SECTION_BG, tcolor=WHITE, halign="center")

ws1.row_dimensions[1].height = 28
ws1.row_dimensions[2].height = 14
ws1.row_dimensions[3].height = 14
ws1.row_dimensions[4].height = 18
ws1.row_dimensions[5].height = 16
ws1.row_dimensions[6].height = 10

# ── Stat boxes ────────────────────────────────────────────────────────────────
section_header(ws1, 7, 1, 12, "  ▸  SUMMARY STATISTICS")
ws1.row_dimensions[7].height = 18

# Row 8-9: February stats, Row 10-11: March stats
stat_box(ws1, 8,  1,  "FEB — Working Days",   20,  "days",  NAVY, WHITE, GOLD)
stat_box(ws1, 8,  3,  "FEB — Total Hours",    156, "hrs",   "1E3A5F", WHITE, GOLD)
stat_box(ws1, 8,  5,  "FEB — Tasks Completed",7,   "tasks", "143D52", WHITE, GOLD)
stat_box(ws1, 8,  7,  "FEB — Projects",       2,   "",      "0F2D3D", WHITE, GOLD)
stat_box(ws1, 8,  9,  "FEB — Status",         "ALL DONE", "", "065F46", WHITE, "D1FAE5")
stat_box(ws1, 8,  11, "FEB — Avg Hrs/Day",    7.8, "hrs",   "1D4ED8", WHITE, GOLD)

stat_box(ws1, 10, 1,  "MAR — Working Days",   20,  "days",  "7C3AED", WHITE, GOLD)
stat_box(ws1, 10, 3,  "MAR — Total Hours",    159, "hrs",   "6D28D9", WHITE, GOLD)
stat_box(ws1, 10, 5,  "MAR — Tasks Completed",18,  "tasks", "5B21B6", WHITE, GOLD)
stat_box(ws1, 10, 7,  "MAR — Projects",       3,   "",      "4C1D95", WHITE, GOLD)
stat_box(ws1, 10, 9,  "MAR — Leaves/Off",     2,   "days",  "92400E", WHITE, GOLD)
stat_box(ws1, 10, 11, "MAR — Avg Hrs/Day",    7.95,"hrs",   "1D4ED8", WHITE, GOLD)

for r in [8, 9, 10, 11]:
    ws1.row_dimensions[r].height = 22
ws1.row_dimensions[12].height = 10

# ── Projects overview ─────────────────────────────────────────────────────────
section_header(ws1, 13, 1, 12, "  ▸  PROJECTS WORKED ON")
ws1.row_dimensions[13].height = 18

col_header(ws1, 14, [
    (1, "Project"), (3, "Type"), (5, "Duration"),
    (7, "Key Deliverable"), (10, "Status"), (12, "Tech Stack")
])
ws1.row_dimensions[14].height = 18

projects = [
    ("Curiosity Hub\n(InvoiceBill)", "College Project\nFull-Stack App",
     "Feb 01 – Mar 21",
     "Invoice/Quotation system, OTP Auth, Notifications,\nPayment Integration (SabPaisa), Security hardening",
     "✅ Completed\n(submitted)", "React Native, Expo, Supabase, Node.js, Railway"),
    ("Orion AI\n(Multi-Agent)", "Personal Learning\nWeb Project",
     "Mar 23 – Mar 27",
     "Multi-LLM agent platform PRD + TRD,\nPrototype research & architecture design",
     "🔄 In Progress", "React, Vite, FastAPI, Groq, OpenAI, Claude"),
    ("Curiosity Hub\n(Web Backend)", "College Project\nBackend",
     "Mar 18 – Mar 20",
     "Architecture planning with Nikunj,\nBackend path design for Curiosity Hub web",
     "🔄 In Progress", "Node.js, Railway, Supabase"),
]

for i, (proj, ptype, dur, deliv, status, tech) in enumerate(projects):
    r = 15 + i
    bg = LIGHT_BG if i % 2 == 0 else ALT_ROW
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell(ws1, r, 1, proj, bold=True, size=9, fcolor=bg, wrap=True, halign="center")
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cell(ws1, r, 3, ptype, size=9, fcolor=bg, wrap=True, halign="center")
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    cell(ws1, r, 5, dur, size=9, fcolor=bg, wrap=True, halign="center")
    ws1.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    cell(ws1, r, 7, deliv, size=9, fcolor=bg, wrap=True)
    ws1.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
    sfg = GREEN_BG if "✅" in status else ORANGE_BG
    stx = GREEN_TXT if "✅" in status else ORANGE_TXT
    cell(ws1, r, 10, status, bold=True, size=9, fcolor=sfg, tcolor=stx,
         wrap=True, halign="center")
    cell(ws1, r, 12, tech, size=9, fcolor=bg, wrap=True)
    ws1.row_dimensions[r].height = 40

ws1.row_dimensions[18].height = 10

# ── Skills summary ────────────────────────────────────────────────────────────
section_header(ws1, 19, 1, 12, "  ▸  KEY SKILLS ACQUIRED / IMPROVED THIS PERIOD")
ws1.row_dimensions[19].height = 18

col_header(ws1, 20, [(1, "Category"), (4, "Skills"), (10, "Level"), (12, "Month")])
ws1.row_dimensions[20].height = 16

skills = [
    ("Mobile Development",  "React Native, Expo Push Notifications, WebSocket integration",         "Advanced",     "Feb–Mar", BLUE_BG, BLUE_TXT),
    ("Backend / Database",  "Node.js, FastAPI (Python), Supabase, PostgreSQL, Railway deployment",  "Intermediate", "Feb–Mar", GREEN_BG, GREEN_TXT),
    ("Security",            "MOBSF testing, VirusTotal, rate limiting, RLS, auth audit logging",    "Intermediate", "March",   ORANGE_BG, ORANGE_TXT),
    ("Payment Systems",     "SabPaisa, Trustopay APK integration, payment error debugging",         "Beginner",     "March",   RED_BG, RED_TXT),
    ("AI / Multi-Agent",    "LangGraph, CrewAI, AutoGen, MCP, LangSmith, Vector DBs, ReAct",       "Learning",     "March",   PURPLE_BG, PURPLE_TXT),
    ("Dev Tools & Workflow","GitHub, Railway CI/CD, MOBSF, openpyxl, PRD/TRD writing",             "Intermediate", "Feb–Mar", ALT_ROW, DARK_TXT),
]

for i, (cat, sk, level, month, bg, tx) in enumerate(skills):
    r = 21 + i
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell(ws1, r, 1, cat, bold=True, size=9, fcolor=bg, tcolor=tx, wrap=True)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    cell(ws1, r, 4, sk, size=9, fcolor=bg, wrap=True)
    ws1.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
    cell(ws1, r, 10, level, bold=True, size=9, fcolor=bg, tcolor=tx, halign="center")
    cell(ws1, r, 12, month, size=9, fcolor=bg, tcolor=tx, halign="center")
    ws1.row_dimensions[r].height = 22

# Column widths
widths = {1:16, 2:16, 3:14, 4:14, 5:14, 6:14, 7:18, 8:18, 9:18,
          10:14, 11:14, 12:22}
for c, w in widths.items():
    ws1.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — FEBRUARY 2026
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📅 February 2026")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A7"

merge_cell(ws2, 1, 1, 2, 10, "FEBRUARY 2026 — MONTHLY WORK REPORT",
           bold=True, size=16, fcolor=NAVY, tcolor=GOLD)
merge_cell(ws2, 3, 1, 3, 10,
           "Project: Curiosity Hub (InvoiceBill)  ·  Total Working Days: 20  ·  Total Hours: ~156  ·  All Tasks: Completed",
           bold=False, size=9, fcolor=SECTION_BG, tcolor=WHITE)
ws2.row_dimensions[1].height = 26
ws2.row_dimensions[2].height = 14
ws2.row_dimensions[3].height = 16
ws2.row_dimensions[4].height = 10

# Stats row
stat_pairs = [
    ("Working Days", "20", "1", "2"),
    ("Total Hours", "~156", "3", "4"),
    ("Tasks Done", "7 Blocks", "5", "6"),
    ("Leaves/Off", "0", "7", "8"),
    ("Avg Hrs/Day", "7.8 hrs", "9", "10"),
]
for label, val, c1, c2 in stat_pairs:
    c1, c2 = int(c1), int(c2)
    ws2.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
    lc = ws2.cell(row=5, column=c1, value=label)
    lc.font = font(bold=True, size=8, color=WHITE)
    lc.fill = fill(NAVY)
    lc.alignment = align(h="center")
    lc.border = border(style="medium", color=GOLD)
    ws2.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    vc = ws2.cell(row=6, column=c1, value=val)
    vc.font = font(bold=True, size=14, color=GOLD)
    vc.fill = fill(SECTION_BG)
    vc.alignment = align(h="center")
    vc.border = border(style="medium", color=GOLD)
ws2.row_dimensions[5].height = 20
ws2.row_dimensions[6].height = 26
ws2.row_dimensions[7].height = 8

section_header(ws2, 8, 1, 10, "  ▸  WEEKLY TASK BREAKDOWN")
ws2.row_dimensions[8].height = 18
col_header(ws2, 9, [
    (1, "Week"), (2, "Date Range"), (3, "Days"),
    (4, "Task Description"), (8, "Technologies"), (9, "Status"), (10, "Hours")
])
ws2.row_dimensions[9].height = 20

feb_weeks = [
    ("Week 1", "01 Feb – 04 Feb", "4",
     "Curiosity Hub: Applied validations in MagicPatterns. Fixed InvoiceBill errors. "
     "Attempted Facebook account access. Created database schema & tables in Supabase. "
     "Worked on OTP verification & template fixes.",
     "Supabase, MagicPatterns, React Native", "✅ Completed", "~30"),
    ("Week 2", "05 Feb – 08 Feb", "4",
     "Implemented email OTP verification system. Created functions and row-level "
     "security policies in Supabase. Resolved data saving issues in Supabase.",
     "Supabase RLS, Email OTP", "✅ Completed", "~28"),
    ("Week 3", "09 Feb – 12 Feb", "4",
     "Implemented invoice data storage in Supabase with UI visibility for seller "
     "and buyer. Fixed fetch issues. Developed contact form with email notification "
     "and admin verification. Started React Native InvoiceBill (PRD & frontend).",
     "Supabase, React Native, Email API", "✅ Completed", "~28"),
    ("Week 4", "13 Feb – 16 Feb", "4",
     "Completed major part of React Native frontend. Learned PostgreSQL basics. "
     "Created backend roadmap and database tables. Started backend setup and explored "
     "Railway deployment with Node.js.",
     "React Native, PostgreSQL, Node.js, Railway", "✅ Completed", "~28"),
    ("Week 5", "17 Feb – 20 Feb", "4",
     "Deployed frontend temporarily. Fixed frontend bugs. Added splash screen and "
     "animations. Resolved GitHub–Railway connection issues. Deployed backend to "
     "Railway. Worked on project documentation.",
     "Railway, GitHub, Node.js, React Native", "✅ Completed", "~24"),
    ("Week 6", "21 Feb – 24 Feb", "4",
     "Worked on Curiosity project. Implemented user registration & login. "
     "Developed invoice & quotation creation features and profile functionality.",
     "React Native, Supabase, Node.js", "✅ Completed", "~24"),
    ("Week 7", "25 Feb – 28 Feb", "4",
     "Implemented notifications system. Fixed system bugs. Improved performance "
     "with sessions. Added onboarding screens. Removed dummy data. Improved UI states. "
     "Implemented: single user login per phone, customer contact import, recurring invoice creation.",
     "Expo Notifications, React Native, Supabase", "✅ Completed", "~24"),
]

for i, (wk, dates, days, task, tech, status, hrs) in enumerate(feb_weeks):
    r = 10 + i
    bg = LIGHT_BG if i % 2 == 0 else ALT_ROW
    cell(ws2, r, 1, wk, bold=True, size=9, fcolor=NAVY, tcolor=GOLD, halign="center")
    cell(ws2, r, 2, dates, size=9, fcolor=bg, halign="center")
    cell(ws2, r, 3, days, size=9, fcolor=bg, halign="center")
    ws2.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    cell(ws2, r, 4, task, size=9, fcolor=bg, wrap=True)
    cell(ws2, r, 8, tech, size=9, fcolor=bg, wrap=True)
    cell(ws2, r, 9, status, bold=True, size=9, fcolor=GREEN_BG, tcolor=GREEN_TXT, halign="center")
    cell(ws2, r, 10, hrs, bold=True, size=9, fcolor=bg, halign="center")
    ws2.row_dimensions[r].height = 55

ws2.row_dimensions[17].height = 10

# Achievements section
section_header(ws2, 18, 1, 10, "  ▸  FEBRUARY KEY ACHIEVEMENTS & LEARNINGS")
ws2.row_dimensions[18].height = 18
col_header(ws2, 19, [(1, "Category"), (4, "Details")], fcolor=SECTION_BG)
ws2.row_dimensions[19].height = 16

feb_highlights = [
    ("🏆 Key Achievement", "Built full InvoiceBill application with OTP auth, invoice/quotation creation, Supabase storage, email notifications, and recurring invoices from scratch to deployment."),
    ("⚙️ Technical Skills", "Mastered Supabase RLS policies, Row-level security, email OTP systems, Railway CI/CD deployment pipeline, React Native animation APIs."),
    ("🗄️ Database", "Designed and implemented full database schema in Supabase for users, invoices, quotations, items, and customers with proper foreign keys."),
    ("🚀 Deployment", "Successfully deployed both frontend (React Native/Expo) and backend (Node.js) to Railway with GitHub CI/CD integration."),
    ("📱 Mobile App", "Delivered production-quality React Native screens: onboarding, login, registration, invoice creation, quotation management, profile."),
]
for i, (cat, detail) in enumerate(feb_highlights):
    r = 20 + i
    bg = LIGHT_BG if i % 2 == 0 else ALT_ROW
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell(ws2, r, 1, cat, bold=True, size=9, fcolor=NAVY, tcolor=GOLD)
    ws2.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
    cell(ws2, r, 4, detail, size=9, fcolor=bg, wrap=True)
    ws2.row_dimensions[r].height = 30

# Column widths
w2 = {1:10, 2:16, 3:8, 4:18, 5:18, 6:18, 7:18, 8:22, 9:14, 10:10}
for c, w in w2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — MARCH 2026 (DAILY LOG)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📅 March 2026")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A8"

merge_cell(ws3, 1, 1, 2, 11, "MARCH 2026 — MONTHLY WORK REPORT (DAILY LOG)",
           bold=True, size=16, fcolor=NAVY, tcolor=GOLD)
merge_cell(ws3, 3, 1, 3, 11,
           "Projects: Curiosity Hub + Orion AI Multi-Agent  ·  Working Days: 20  ·  Total Hours: 159  ·  Period: 02 Mar – 27 Mar 2026",
           bold=False, size=9, fcolor=SECTION_BG, tcolor=WHITE)
ws3.row_dimensions[1].height = 26
ws3.row_dimensions[2].height = 14
ws3.row_dimensions[3].height = 16
ws3.row_dimensions[4].height = 10

stat_pairs3 = [
    ("Working Days", "20"),
    ("Total Hours", "159 hrs"),
    ("Leave/Off Days", "2"),
    ("Avg Hrs/Day", "7.95 hrs"),
    ("Projects Active", "3"),
]
cols3 = [(1,2), (3,4), (5,6), (7,8), (9,10)]
for (label, val), (c1, c2) in zip(stat_pairs3, cols3):
    ws3.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
    lc = ws3.cell(row=5, column=c1, value=label)
    lc.font = font(bold=True, size=8, color=WHITE)
    lc.fill = fill(NAVY)
    lc.alignment = align(h="center")
    lc.border = border(style="medium", color=GOLD)
    ws3.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    vc = ws3.cell(row=6, column=c1, value=val)
    vc.font = font(bold=True, size=13, color=GOLD)
    vc.fill = fill(SECTION_BG)
    vc.alignment = align(h="center")
    vc.border = border(style="medium", color=GOLD)
ws3.row_dimensions[5].height = 20
ws3.row_dimensions[6].height = 24
ws3.row_dimensions[7].height = 8

section_header(ws3, 8, 1, 11, "  ▸  DAILY WORK LOG")
ws3.row_dimensions[8].height = 18
col_header(ws3, 9, [
    (1, "#"), (2, "Date"), (3, "Day"),
    (4, "Task / Work Done"), (8, "Project"), (9, "Hours"),
    (10, "Start"), (11, "End")
])
ws3.row_dimensions[9].height = 18

# Status color logic
def row_style(status):
    if status == "leave":    return RED_BG,    RED_TXT
    if status == "college":  return ORANGE_BG, ORANGE_TXT
    if status == "learn":    return PURPLE_BG, PURPLE_TXT
    return LIGHT_BG, DARK_TXT

march_log = [
    # (row#, date, day, task, project, hours, start, end, status)
    (37, "02-03-2026", "Monday",
     "Transactions (Add money, Transfer, Withdrawal) fully integrated. "
     "Implemented settle-back gesture interaction. Prevented direct jump to Home page without login.",
     "Curiosity Hub", 10, "10:00", "08:00", "work"),
    (38, "03-03-2026", "Tuesday",
     "App crashes immediately after launch — intensive debugging session. "
     "Reverted problematic code. Merged one module at a time to isolate crash. "
     "All requirements list completed. Business Profile, Invoice Settings, Terms & Conditions, "
     "Verification, My Addresses, Item List data saved to Database.",
     "Curiosity Hub", 10, "10:00", "08:00", "work"),
    (39, "04-03-2026", "Wednesday",
     "Invoice/Quotation/Recurring notifications now fire via Expo when created or reminder sent. "
     "Implemented inbuilt Notification Center. Fixed customer-disappear bug that occurred after ~1 hour.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (40, "05-03-2026", "Thursday",
     "Integrated Forgot Password feature end-to-end. Researched SabPaisa payment platform "
     "and its role in the app. Fixed remaining bugs in the system. "
     "Fixed Curiosity Hub Application frontend issues.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (41, "06-03-2026", "Friday",
     "Fixed critical security bug: direct jump to Home page possible without being logged in. "
     "Fixed other minor bugs. Addressed frontend vulnerability issues in the system.",
     "Curiosity Hub", 7, "10:00", "05:00", "work"),
    (42, "07-03-2026", "Saturday",
     "College reporting day — attended college for project status presentation.",
     "College", 0, "—", "—", "college"),
    (43, "09-03-2026", "Sunday/Mon",
     "Built system to deliver invoices/quotations even when User B is inactive. "
     "When User B signs in, they can view all old and new notifications, invoices, and quotations.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (44, "10-03-2026", "Tuesday",
     "Integrated WebSocket into the application for real-time updates. "
     "Studied SabPaisa payment gateway. Learned advanced useEffect patterns, "
     "Global State management to bring the app to industry level.",
     "Curiosity Hub", 7, "10:00", "05:00", "work"),
    (45, "11-03-2026", "Wednesday",
     "Integrated useEffect and Global State for smoother app experience. "
     "Integrated SabPaisa transaction management system. "
     "Fixed direct-login bug precisely — user no longer bypasses auth.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (46, "12-03-2026", "Thursday",
     "On Leave.",
     "—", 0, "—", "—", "leave"),
    (47, "13-03-2026", "Friday",
     "Push Notifications now working consistently across all notification types. "
     "Improved sender and receiver system architecture for better reliability.",
     "Curiosity Hub", 7, "10:00", "05:00", "work"),
    (48, "16-03-2026", "Monday",
     "Users can now upload and update profile images (Supabase Storage bucket integrated). "
     "Researched and learned mobile security testing tools. "
     "Tested app in MOBSF tool — Security Score: 55/100. Implemented improvements.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (49, "17-03-2026", "Tuesday",
     "Debugged errors occurring during production APK creation. "
     "Integrated SabPaisa Test credentials into payment system. "
     "Attempting to solve 'Client code is invalid' payment error.",
     "Curiosity Hub", 7, "10:00", "05:00", "work"),
    (50, "18-03-2026", "Wednesday",
     "Profile image upload & update fully functional. Continued debugging "
     "'Client id is invalid' payment error. Collaborated with Nikunj to plan "
     "architecture and further backend development for Curiosity Hub web.",
     "Curiosity Hub / Curiosity Web", 8, "10:00", "06:00", "work"),
    (51, "19-03-2026", "Thursday",
     "Tested Trustopay APK on MOBSF tool and VirusTotal. "
     "Security identified as the current weak point — planning improvements.",
     "Curiosity Hub", 7, "10:00", "05:00", "work"),
    (52, "20-03-2026", "Friday",
     "Continued debugging payment system ('Client code is invalid' error persists). "
     "Built further backend architecture for Curiosity Hub web with Nikunj.",
     "Curiosity Hub / Curiosity Web", 8, "10:00", "06:00", "work"),
    (53, "21-03-2026", "Saturday",
     "Implemented comprehensive security improvements: "
     "(1) Secure authentication — audit logging for login, registration, password reset, token refresh. "
     "(2) Prevented unauthorized data access — Supabase via HTTPS only, per-user auth checks. "
     "(3) Secure deployment — HTTPS enforcement, Helmet headers, audit logs. "
     "(4) Bot/abuse prevention — rate limiting (5/15min auth, 100/min API). "
     "(5) Secrets protection — env vars, .env excluded from Git. "
     "(6) Input validation — DTOs with class-validator, UUID checks, safe avatar uploads.",
     "Curiosity Hub", 8, "10:00", "06:00", "work"),
    (54, "23-03-2026 to\n24-03-2026", "Mon–Tue",
     "Deep learning session on AI Agent Architecture: "
     "Architectural Patterns (ReAct, Multi-Agent Orchestration, Task Decomposition, "
     "Self-Reflection & Self-Correction Loops, Plan-and-Execute). "
     "Frameworks (LangGraph, CrewAI, Microsoft AutoGen, PydanticAI, LlamaIndex). "
     "Core Skills (Function Calling/Tool Use, MCP, Memory Management, State Persistence, HITL). "
     "Infrastructure (LangSmith, AgentOps, Braintrust, Guardrails AI, Vector Databases). "
     "Evaluation (LLM-as-a-Judge, Unit Testing for Tools, Agentic Benchmarking).",
     "Self Learning", 4, "10:00", "02:00", "learn"),
    (55, "26-03-2026", "Thursday",
     "Researched how to build Multi AI Agent projects. "
     "Developed a working prototype version of Orion AI. "
     "Researched Product Requirements Document (PRD) and Technical Requirements Document (TRD) formats.",
     "Orion AI", 8, "11:00", "07:00", "work"),
    (56, "27-03-2026", "Friday",
     "Creating comprehensive PRD and TRD documents for Orion AI Multi-Agent Platform. "
     "Generated professional PDF versions using Python + ReportLab. "
     "Full system architecture, database schema, API design, and 7-day dev checklist completed.",
     "Orion AI", 8, "10:00", "06:00", "work"),
]

for i, (row_num, date, day, task, project, hours, start, end, status) in enumerate(march_log):
    r = 10 + i
    bg, tx = row_style(status)
    alt_bg = ALT_ROW if i % 2 else bg

    cell(ws3, r, 1, row_num, bold=True, size=8, fcolor=NAVY, tcolor=GOLD, halign="center")
    cell(ws3, r, 2, date, size=8, fcolor=bg, halign="center")
    cell(ws3, r, 3, day, size=8, fcolor=bg, halign="center")
    ws3.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    cell(ws3, r, 4, task, size=8.5, fcolor=bg, wrap=True)
    cell(ws3, r, 8, project, bold=True, size=8, fcolor=bg, tcolor=tx, wrap=True, halign="center")

    if hours == 0:
        cell(ws3, r, 9, "—", bold=True, size=9, fcolor=RED_BG, tcolor=RED_TXT, halign="center")
    else:
        cell(ws3, r, 9, hours, bold=True, size=9, fcolor=bg, halign="center")
    cell(ws3, r, 10, start, size=8, fcolor=bg, halign="center")
    cell(ws3, r, 11, end, size=8, fcolor=bg, halign="center")
    ws3.row_dimensions[r].height = 55

ws3.row_dimensions[30].height = 10

# March achievements
section_header(ws3, 31, 1, 11, "  ▸  MARCH KEY ACHIEVEMENTS, LEARNINGS & CHALLENGES")
ws3.row_dimensions[31].height = 18
col_header(ws3, 32, [(1, "Category"), (3, "Details")], fcolor=SECTION_BG)
ws3.row_dimensions[32].height = 16

mar_highlights = [
    ("🏆 Top Achievement",
     "Delivered Curiosity Hub mobile app with full transaction management (Add/Transfer/Withdraw), "
     "push notifications, real-time WebSocket, payment gateway integration (SabPaisa), "
     "profile image uploads, and security hardening — all in one month.",
     GREEN_BG, GREEN_TXT),
    ("🔒 Security Work",
     "Ran MOBSF security audit (score: 55), tested on VirusTotal, implemented rate limiting, "
     "auth audit logging, Helmet headers, RLS, input validation (DTOs), and env secret management.",
     ORANGE_BG, ORANGE_TXT),
    ("💳 Payment Integration",
     "Integrated SabPaisa and Trustopay payment systems. Encountered 'Client code is invalid' "
     "error — researched extensively, integration is partially working (test mode).",
     RED_BG, RED_TXT),
    ("🤖 AI Agent Learning",
     "Completed deep-dive on Multi-Agent AI architecture: LangGraph, CrewAI, AutoGen, MCP, "
     "LangSmith, Vector DBs, ReAct pattern, HITL, LLM-as-a-Judge evaluation.",
     PURPLE_BG, PURPLE_TXT),
    ("🚀 New Project Launched",
     "Conceived, researched, and started Orion AI — a multi-LLM document generation platform "
     "using Groq, OpenAI GPT-4o, and Claude Sonnet 4.6. PRD + TRD fully completed.",
     BLUE_BG, BLUE_TXT),
    ("⚠️ Main Challenge",
     "SabPaisa payment 'Client code is invalid' error remained unresolved by month end. "
     "App crash after launch (row 38) required full debugging and code reversal — resolved same day.",
     ORANGE_BG, ORANGE_TXT),
]
for i, (cat, detail, bg, tx) in enumerate(mar_highlights):
    r = 33 + i
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell(ws3, r, 1, cat, bold=True, size=9, fcolor=bg, tcolor=tx)
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    cell(ws3, r, 3, detail, size=9, fcolor=LIGHT_BG if i % 2 else ALT_ROW, wrap=True)
    ws3.row_dimensions[r].height = 35

# Column widths
w3 = {1:6, 2:14, 3:10, 4:18, 5:18, 6:18, 7:18, 8:16, 9:8, 10:8, 11:8}
for c, w in w3.items():
    ws3.column_dimensions[get_column_letter(c)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — HOURS TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("⏱️ Hours Tracker")
ws4.sheet_view.showGridLines = False

merge_cell(ws4, 1, 1, 2, 8, "HOURS TRACKER — March 2026 Daily Breakdown",
           bold=True, size=14, fcolor=NAVY, tcolor=GOLD)
ws4.row_dimensions[1].height = 24
ws4.row_dimensions[2].height = 12
ws4.row_dimensions[3].height = 8

col_header(ws4, 4, [
    (1, "Row #"), (2, "Date"), (3, "Day"), (4, "Project"),
    (5, "Hours"), (6, "Start"), (7, "End"), (8, "Type")
])
ws4.row_dimensions[4].height = 18

hours_data = [
    (37, "02-03-2026", "Monday",    "Curiosity Hub",      10, "10:00", "08:00", "Work"),
    (38, "03-03-2026", "Tuesday",   "Curiosity Hub",      10, "10:00", "08:00", "Work"),
    (39, "04-03-2026", "Wednesday", "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (40, "05-03-2026", "Thursday",  "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (41, "06-03-2026", "Friday",    "Curiosity Hub",       7, "10:00", "05:00", "Work"),
    (42, "07-03-2026", "Saturday",  "College",             0, "—",     "—",     "College"),
    (43, "09-03-2026", "Sunday",    "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (44, "10-03-2026", "Tuesday",   "Curiosity Hub",       7, "10:00", "05:00", "Work"),
    (45, "11-03-2026", "Wednesday", "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (46, "12-03-2026", "Thursday",  "—",                   0, "—",     "—",     "Leave"),
    (47, "13-03-2026", "Friday",    "Curiosity Hub",       7, "10:00", "05:00", "Work"),
    (48, "16-03-2026", "Monday",    "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (49, "17-03-2026", "Tuesday",   "Curiosity Hub",       7, "10:00", "05:00", "Work"),
    (50, "18-03-2026", "Wednesday", "Curiosity Hub/Web",   8, "10:00", "06:00", "Work"),
    (51, "19-03-2026", "Thursday",  "Curiosity Hub",       7, "10:00", "05:00", "Work"),
    (52, "20-03-2026", "Friday",    "Curiosity Hub/Web",   8, "10:00", "06:00", "Work"),
    (53, "21-03-2026", "Saturday",  "Curiosity Hub",       8, "10:00", "06:00", "Work"),
    (54, "23-03-2026", "Monday",    "Self Learning",        2, "10:00", "12:00", "Learning"),
    (55, "24-03-2026", "Tuesday",   "Self Learning",        2, "10:00", "12:00", "Learning"),
    (56, "26-03-2026", "Thursday",  "Orion AI",            8, "11:00", "07:00", "Work"),
    (57, "27-03-2026", "Friday",    "Orion AI",            8, "10:00", "06:00", "Work"),
]

type_colors = {
    "Work":     (LIGHT_BG,  DARK_TXT),
    "Leave":    (RED_BG,    RED_TXT),
    "College":  (ORANGE_BG, ORANGE_TXT),
    "Learning": (PURPLE_BG, PURPLE_TXT),
}

for i, (rn, date, day, proj, hrs, start, end, typ) in enumerate(hours_data):
    r = 5 + i
    bg, tx = type_colors.get(typ, (LIGHT_BG, DARK_TXT))
    cell(ws4, r, 1, rn,    size=9, fcolor=NAVY, tcolor=GOLD, halign="center", bold=True)
    cell(ws4, r, 2, date,  size=9, fcolor=bg, halign="center")
    cell(ws4, r, 3, day,   size=9, fcolor=bg, halign="center")
    cell(ws4, r, 4, proj,  size=9, fcolor=bg)
    cell(ws4, r, 5, hrs,   size=9, fcolor=bg, halign="center", bold=(hrs>0))
    cell(ws4, r, 6, start, size=9, fcolor=bg, halign="center")
    cell(ws4, r, 7, end,   size=9, fcolor=bg, halign="center")
    cell(ws4, r, 8, typ,   size=9, fcolor=bg, tcolor=tx, halign="center", bold=True)
    ws4.row_dimensions[r].height = 16

# Totals
r_total = 5 + len(hours_data)
ws4.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=4)
tc = ws4.cell(row=r_total, column=1, value="TOTAL HOURS (MARCH 2026)")
tc.font = font(bold=True, size=10, color=WHITE)
tc.fill = fill(NAVY)
tc.alignment = align(h="right")
tc.border = border(style="medium", color=GOLD)
hrs_cell = ws4.cell(row=r_total, column=5, value=f"=SUM(E5:E{r_total-1})")
hrs_cell.font = font(bold=True, size=12, color=GOLD)
hrs_cell.fill = fill(NAVY)
hrs_cell.alignment = align(h="center")
hrs_cell.border = border(style="medium", color=GOLD)
ws4.row_dimensions[r_total].height = 22

w4 = {1:8, 2:14, 3:12, 4:22, 5:10, 6:10, 7:10, 8:12}
for c, w in w4.items():
    ws4.column_dimensions[get_column_letter(c)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"Monthly report saved to:\n   {OUTPUT}")
